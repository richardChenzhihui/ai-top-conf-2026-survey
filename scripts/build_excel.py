"""
Build the 5-sheet survey Excel from the three classified JSON files.

Usage:
  python3 scripts/build_excel.py

Output:
  survey_2026.xlsx   (in the project root)
"""

import json
import re
from pathlib import Path
from collections import Counter, defaultdict

try:
    import openpyxl
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side
    )
    from openpyxl.chart import BarChart, Reference
    from openpyxl.utils import get_column_letter
except ImportError:
    import subprocess, sys
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.chart import BarChart, Reference
    from openpyxl.utils import get_column_letter

BASE = Path(__file__).parent.parent
ANA  = BASE / "data" / "_analysis"
OUT  = BASE / "survey_2026.xlsx"

CONFS = ["cvpr", "iclr", "neurips"]
CONF_LABELS = {"cvpr": "CVPR 2026", "iclr": "ICLR 2026", "neurips": "NeurIPS 2025"}

CORE_CATS = [
    "多模态大模型（MLLM/VLM）",
    "强化学习后训练（RLHF/DPO/GRPO/奖励模型）",
    "Agent 系统",
    "Agent 后训练",
    "大语言模型预训练（架构/数据/规模）",
    "大模型 Infra（推理/训练/效率系统）",
]

ALL_CATS = [
    "多模态大模型（MLLM/VLM）",
    "强化学习后训练（RLHF/DPO/GRPO/奖励模型）",
    "Agent 系统",
    "Agent 后训练",
    "大语言模型预训练（架构/数据/规模）",
    "大模型 Infra（推理/训练/效率系统）",
    "推理与规划（CoT/数学/逻辑，非 RL 类）",
    "代码生成与编程语言处理",
    "安全与对齐（幻觉/偏见/毒性/越狱，非 RL 类）",
    "评测与基准构建",
    "图神经网络与结构化学习",
    "理论与优化",
    "对话系统与文本生成",
    "目标检测与分割",
    "生成模型（扩散/GAN/视频生成/图像编辑，非 MLLM 驱动）",
    "三维视觉（NeRF/3DGS/深度估计/点云）",
    "视频理解（动作识别/跟踪/时序建模，非 MLLM 类）",
    "底层视觉（超分/去噪/增强）",
    "人体相关（姿态/人脸/手势/运动合成）",
    "医学影像（非多模态大模型类）",
    "自动驾驶（感知/预测/规划）",
    "遥感与卫星",
    "其他",
]

# ---------------------------------------------------------------------------
# Style helpers
# ---------------------------------------------------------------------------
HEADER_FILL  = PatternFill("solid", fgColor="1F3864")
HEADER_FONT  = Font(bold=True, color="FFFFFF", size=10)
SUBHEAD_FILL = PatternFill("solid", fgColor="2E75B6")
SUBHEAD_FONT = Font(bold=True, color="FFFFFF", size=10)
ALT_FILL_A   = PatternFill("solid", fgColor="EBF3FB")
ALT_FILL_B   = PatternFill("solid", fgColor="FFFFFF")
BORDER_THIN  = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center")

def header_cell(ws, row, col, value, fill=None):
    c = ws.cell(row=row, column=col, value=value)
    c.fill  = fill or HEADER_FILL
    c.font  = HEADER_FONT
    c.alignment = CENTER
    c.border = BORDER_THIN
    return c

_ILLEGAL = re.compile(r'[\x00-\x08\x0b\x0c\x0e-\x1f]')

def _clean(val):
    if isinstance(val, str):
        return _ILLEGAL.sub(' ', val)
    return val

def data_cell(ws, row, col, value, alt=False, wrap=True):
    c = ws.cell(row=row, column=col, value=_clean(value))
    c.fill  = ALT_FILL_A if alt else ALT_FILL_B
    c.alignment = WRAP if wrap else Alignment(vertical="top")
    c.border = BORDER_THIN
    return c

def freeze_and_filter(ws, row=2, col=1):
    ws.freeze_panes = ws.cell(row=row, column=col)
    ws.auto_filter.ref = ws.dimensions

# ---------------------------------------------------------------------------
# Load data
# ---------------------------------------------------------------------------
def load_all():
    all_papers = []
    per_conf = {}
    for conf in CONFS:
        path = ANA / f"{conf}_classified.json"
        if not path.exists():
            print(f"[WARN] {path} not found, skipping")
            per_conf[conf] = []
            continue
        papers = json.loads(path.read_text())
        for p in papers:
            p["_conf_key"] = conf
        per_conf[conf] = papers
        all_papers.extend(papers)
    return all_papers, per_conf

# ---------------------------------------------------------------------------
# Sheet 1: 全量列表
# ---------------------------------------------------------------------------
COLS_ALL = [
    ("序号",        "idx",          5),
    ("会议",        "conference",   8),
    ("Track",       "track",        9),
    ("论文标题",    "title_en",     55),
    ("L1 类别",     "category_l1",  22),
    ("L2 子方向",   "category_l2",  18),
    ("关键词",      "keywords",     25),
    ("摘要",        "abstract",     60),
]

def build_all_sheet(wb, papers):
    ws = wb.create_sheet("全量列表")
    headers = [c[0] for c in COLS_ALL]
    widths  = [c[2] for c in COLS_ALL]
    fields  = [c[1] for c in COLS_ALL]

    for col, h in enumerate(headers, 1):
        header_cell(ws, 1, col, h)
    ws.row_dimensions[1].height = 22

    # Sort: conference order → L1 → L2
    conf_order = {c: i for i, c in enumerate(CONFS)}
    sorted_papers = sorted(
        papers,
        key=lambda p: (
            conf_order.get(p.get("_conf_key", ""), 99),
            ALL_CATS.index(p.get("category_l1", "其他")) if p.get("category_l1", "其他") in ALL_CATS else 99,
            p.get("category_l2", ""),
        )
    )

    for row_idx, p in enumerate(sorted_papers, 2):
        alt = row_idx % 2 == 0
        for col, field in enumerate(fields, 1):
            val = p.get(field, "")
            if field == "idx":
                val = row_idx - 1
            data_cell(ws, row_idx, col, val, alt=alt, wrap=(field in ("title_en", "abstract", "keywords")))

    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    freeze_and_filter(ws)
    ws.sheet_view.showGridLines = False
    print(f"[Sheet 1] 全量列表: {len(sorted_papers)} rows")

# ---------------------------------------------------------------------------
# Sheet 2: 核心热点统计
# ---------------------------------------------------------------------------
def build_core_sheet(wb, per_conf):
    ws = wb.create_sheet("核心热点统计")

    # --- Table 1: Core L1 × Conference ---
    header_cell(ws, 1, 1, "核心类别")
    header_cell(ws, 1, 2, "CVPR 2026")
    header_cell(ws, 1, 3, "ICLR 2026")
    header_cell(ws, 1, 4, "NeurIPS 2025")
    header_cell(ws, 1, 5, "三会总计")

    core_counts = {}
    for cat in CORE_CATS:
        row_data = {}
        for conf in CONFS:
            papers = per_conf.get(conf, [])
            row_data[conf] = sum(1 for p in papers if p.get("category_l1") == cat)
        core_counts[cat] = row_data

    for r, cat in enumerate(CORE_CATS, 2):
        alt = r % 2 == 0
        data_cell(ws, r, 1, cat,              alt=alt)
        data_cell(ws, r, 2, core_counts[cat]["cvpr"],    alt=alt, wrap=False)
        data_cell(ws, r, 3, core_counts[cat]["iclr"],    alt=alt, wrap=False)
        data_cell(ws, r, 4, core_counts[cat]["neurips"], alt=alt, wrap=False)
        total = sum(core_counts[cat].values())
        data_cell(ws, r, 5, total, alt=alt, wrap=False)

    ws.column_dimensions["A"].width = 30
    for col in ["B", "C", "D", "E"]:
        ws.column_dimensions[col].width = 14

    # --- Chart: Core categories bar chart ---
    chart = BarChart()
    chart.type  = "col"
    chart.grouping = "clustered"
    chart.title = "核心热点类别 — 三会对比"
    chart.y_axis.title = "论文数量"
    chart.x_axis.title = "类别"
    chart.style = 10
    chart.width  = 26
    chart.height = 16

    data_ref = Reference(ws, min_col=2, max_col=4, min_row=1, max_row=1 + len(CORE_CATS))
    cats_ref = Reference(ws, min_col=1, min_row=2, max_row=1 + len(CORE_CATS))
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.series[0].title.v = "CVPR 2026"
    chart.series[1].title.v = "ICLR 2026"
    chart.series[2].title.v = "NeurIPS 2025"

    ws.add_chart(chart, "G1")

    # --- Table 2: L2 breakdown for each core category ---
    start_row = 2 + len(CORE_CATS) + 2
    header_cell(ws, start_row, 1, "L2 子方向详细分布", fill=SUBHEAD_FILL)
    header_cell(ws, start_row, 2, "CVPR 2026",          fill=SUBHEAD_FILL)
    header_cell(ws, start_row, 3, "ICLR 2026",          fill=SUBHEAD_FILL)
    header_cell(ws, start_row, 4, "NeurIPS 2025",       fill=SUBHEAD_FILL)
    header_cell(ws, start_row, 5, "合计",               fill=SUBHEAD_FILL)

    r = start_row + 1
    for cat in CORE_CATS:
        # section header
        c = ws.cell(row=r, column=1, value=f"▶ {cat}")
        c.fill = PatternFill("solid", fgColor="D6E4F0")
        c.font = Font(bold=True, size=9)
        c.alignment = WRAP
        c.border = BORDER_THIN
        for col in range(2, 6):
            cc = ws.cell(row=r, column=col)
            cc.fill = PatternFill("solid", fgColor="D6E4F0")
            cc.border = BORDER_THIN
        r += 1

        # collect L2 distribution
        l2_counts = defaultdict(lambda: {"cvpr": 0, "iclr": 0, "neurips": 0})
        for conf in CONFS:
            for p in per_conf.get(conf, []):
                if p.get("category_l1") == cat:
                    l2 = p.get("category_l2") or "未分类"
                    l2_counts[l2][conf] += 1

        for l2, counts in sorted(l2_counts.items(), key=lambda x: -sum(x[1].values())):
            alt = r % 2 == 0
            data_cell(ws, r, 1, f"  {l2}", alt=alt)
            data_cell(ws, r, 2, counts["cvpr"],    alt=alt, wrap=False)
            data_cell(ws, r, 3, counts["iclr"],    alt=alt, wrap=False)
            data_cell(ws, r, 4, counts["neurips"], alt=alt, wrap=False)
            data_cell(ws, r, 5, sum(counts.values()), alt=alt, wrap=False)
            r += 1

    ws.sheet_view.showGridLines = False
    print("[Sheet 2] 核心热点统计 done")

# ---------------------------------------------------------------------------
# Sheet 3/4/5: Per-conference L1 distribution
# ---------------------------------------------------------------------------
def build_conf_sheet(wb, conf, papers):
    label = CONF_LABELS[conf]
    ws = wb.create_sheet(label)

    total = len(papers)
    l1_count = Counter(p.get("category_l1", "其他") for p in papers)

    # Header row
    header_cell(ws, 1, 1, "L1 类别")
    header_cell(ws, 1, 2, "论文数")
    header_cell(ws, 1, 3, "占比(%)")
    header_cell(ws, 1, 4, "Top-3 L2 子方向")

    for r, cat in enumerate(ALL_CATS, 2):
        cnt = l1_count.get(cat, 0)
        pct = f"{cnt/total*100:.1f}" if total else "0.0"
        alt = r % 2 == 0
        data_cell(ws, r, 1, cat, alt=alt)
        data_cell(ws, r, 2, cnt, alt=alt, wrap=False)
        data_cell(ws, r, 3, float(pct), alt=alt, wrap=False)

        # Top-3 L2
        cat_papers = [p for p in papers if p.get("category_l1") == cat]
        l2s = Counter(p.get("category_l2") or "未分类" for p in cat_papers if p.get("category_l2"))
        top3 = ", ".join(f"{k}({v})" for k, v in l2s.most_common(3))
        data_cell(ws, r, 4, top3, alt=alt)

    # Column widths
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 50

    # Bar chart
    chart = BarChart()
    chart.type    = "bar"
    chart.title   = f"{label} — L1 类别论文数"
    chart.y_axis.title = "论文数量"
    chart.x_axis.title = "类别"
    chart.style   = 10
    chart.width   = 18
    chart.height  = 22

    data_ref = Reference(ws, min_col=2, max_col=2, min_row=1, max_row=1 + len(ALL_CATS))
    cats_ref = Reference(ws, min_col=1, min_row=2, max_row=1 + len(ALL_CATS))
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)

    ws.add_chart(chart, "F1")
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False
    print(f"[Sheet] {label}: {total} papers, {len(l1_count)} categories")

# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    all_papers, per_conf = load_all()
    if not all_papers:
        print("No classified data found. Run classify_keywords.py first.")
        return

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    build_all_sheet(wb, all_papers)
    build_core_sheet(wb, per_conf)
    for conf in CONFS:
        if per_conf.get(conf):
            build_conf_sheet(wb, conf, per_conf[conf])

    wb.save(OUT)
    print(f"\nSaved: {OUT}  ({OUT.stat().st_size//1024} KB)")

if __name__ == "__main__":
    main()
